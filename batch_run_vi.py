"""Batch Vietnamese Video Dubbing — Read URLs from Excel, process each, update status.

Usage:
    python batch_run_vi.py                                  # Use default output/video_link.xlsx
    python batch_run_vi.py --excel path/to/links.xlsx       # Custom Excel path
    python batch_run_vi.py --voice male                     # Skip voice prompt, use male
    python batch_run_vi.py --voice female                   # Skip voice prompt, use female
"""
import argparse
import os
import sys

from openpyxl import load_workbook

import config
from src.utils import setup_logging
from pipeline_vi import run_pipeline_vi, _ask_voice_gender, _get_default_vi_output_dir

logger = setup_logging("batch_run_vi")

DEFAULT_EXCEL = "output/video_link.xlsx"


def main():
    parser = argparse.ArgumentParser(description="Batch Vietnamese Video Dubbing: process URLs from Excel")
    parser.add_argument(
        "--excel",
        default=DEFAULT_EXCEL,
        help=f"Path to Excel file with video links (default: {DEFAULT_EXCEL})",
    )
    parser.add_argument(
        "--source-lang",
        default=config.DEFAULT_SOURCE_LANG,
        help=f"Source language for all videos (default: {config.DEFAULT_SOURCE_LANG})",
    )
    parser.add_argument(
        "--voice",
        choices=["male", "female"],
        default=None,
        help="Voice gender: male or female (if not set, will ask interactively)",
    )
    parser.add_argument(
        "--skip-video",
        action="store_true",
        help="Skip final video merge",
    )
    parser.add_argument(
        "--output-dir",
        default=_get_default_vi_output_dir(),
        help="Output directory (default: ANKO Project/VN)",
    )
    args = parser.parse_args()

    # Resolve voice ID (backend-aware via config.vi_voice; honours TTS_BACKEND_VI)
    if args.voice == "male":
        voice_id = config.vi_voice("male")
    elif args.voice == "female":
        voice_id = config.vi_voice("female")
    else:
        voice_id = _ask_voice_gender()

    # Load Excel
    if not os.path.exists(args.excel):
        logger.error(f"Excel file not found: {args.excel}")
        sys.exit(1)

    wb = load_workbook(args.excel)
    ws = wb.active

    # Find rows that need processing (status column is empty)
    pending_rows = []
    for row_idx in range(2, ws.max_row + 1):
        url = ws.cell(row=row_idx, column=1).value
        status = ws.cell(row=row_idx, column=2).value
        if url and not status:
            pending_rows.append(row_idx)

    if not pending_rows:
        logger.info("No pending videos. All URLs already have a status.")
        return

    logger.info(f"Found {len(pending_rows)} pending video(s) to process (Vietnamese)")
    logger.info("=" * 60)

    success_count = 0
    fail_count = 0

    for i, row_idx in enumerate(pending_rows):
        url = ws.cell(row=row_idx, column=1).value
        logger.info(f"[{i + 1}/{len(pending_rows)}] Processing: {url}")

        try:
            report = run_pipeline_vi(
                url=url,
                file_path=None,
                source_lang=args.source_lang,
                voice_id=voice_id,
                skip_video=args.skip_video,
                output_dir=args.output_dir,
            )

            folder_name = report["session_id"]

            ws.cell(row=row_idx, column=2, value="SUCCESS")
            ws.cell(row=row_idx, column=3, value=folder_name)

            success_count += 1
            logger.info(f"[{i + 1}/{len(pending_rows)}] SUCCESS → {folder_name}")

        except Exception as e:
            error_msg = str(e)[:100]

            ws.cell(row=row_idx, column=2, value=f"FAILED: {error_msg}")
            ws.cell(row=row_idx, column=3, value="")

            fail_count += 1
            logger.error(f"[{i + 1}/{len(pending_rows)}] FAILED: {error_msg}")

        # Save Excel after each video (crash-safe)
        wb.save(args.excel)

    # Summary
    logger.info("=" * 60)
    logger.info("BATCH COMPLETE (Vietnamese)")
    logger.info(f"  Total:   {len(pending_rows)}")
    logger.info(f"  Success: {success_count}")
    logger.info(f"  Failed:  {fail_count}")
    logger.info(f"  Excel:   {args.excel}")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
