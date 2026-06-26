"""Tab: batch-dub a list of videos from JSON (or Excel), AI-translated.

The manual translate gate can't scale across many videos, so batch always uses
AI auto-translation (or a transcript_vi.json that already exists in the work dir).
Status is written back to the source file after each video (crash-safe for JSON),
mirroring batch_run_json.py's schema.
"""
import json
import os
import queue
import tempfile
import threading
import time
from tkinter import filedialog

import customtkinter as ctk

import config
from pipeline_vi import run_pipeline_vi, _get_default_vi_output_dir, LANG_MAP
from src.translator import translate_transcript
from src.translator_claude import translate_via_claude_cli
from src.translator_opencode import translate_via_opencode_cli
from gui.translate_gate import TRANSLATE_MODES

STATUS_COLORS = {
    "waiting": "gray60",
    "processing": "#d29922",
    "success": "#2ea043",
    "failed": "#f85149",
}


class BatchTab(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="transparent")
        self.videos: list[dict] = []
        self.path: str | None = None
        self._rows: dict = {}
        self._ui_queue: "queue.Queue" = queue.Queue()
        self._running = False

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        bar = ctk.CTkFrame(self, fg_color="transparent")
        bar.grid(row=0, column=0, sticky="ew", padx=4, pady=(4, 8))
        ctk.CTkButton(bar, text="Mở danh sách (JSON/Excel)…",
                      command=self._load).pack(side="left")
        ctk.CTkButton(bar, text="Thêm file MP4…",
                      command=self._add_files).pack(side="left", padx=8)
        ctk.CTkButton(bar, text="Thêm từ YouTube…",
                      command=self._open_youtube_dialog).pack(side="left", padx=(0, 8))
        self.start_btn = ctk.CTkButton(bar, text="▶ Bắt đầu", command=self._start,
                                       state="disabled")
        self.start_btn.pack(side="left", padx=8)
        ctk.CTkLabel(bar, text="Ngôn ngữ:").pack(side="left", padx=(16, 4))
        self.lang_var = ctk.StringVar(value=config.DEFAULT_SOURCE_LANG)
        ctk.CTkOptionMenu(bar, width=110, variable=self.lang_var,
                          values=["en", "ja", "zh", "en-US", "ja-JP", "zh-CN"]).pack(side="left")
        self.skip_video_var = ctk.BooleanVar(value=False)  # mặc định: đóng gói (ghép video)
        ctk.CTkCheckBox(bar, text="Bỏ qua ghép video (chỉ tạo audio)",
                        variable=self.skip_video_var).pack(side="left", padx=(16, 0))
        ctk.CTkLabel(bar, text="Âm nền:").pack(side="left", padx=(16, 4))
        self.bg_var = ctk.StringVar(value="duck")
        ctk.CTkOptionMenu(bar, width=90, variable=self.bg_var,
                          values=["duck", "demucs", "none"]).pack(side="left")
        ctk.CTkLabel(bar, text="Dịch:").pack(side="left", padx=(16, 4))
        self.tmode_var = ctk.StringVar(value="Opencode tự động (đọc & dịch)")
        ctk.CTkOptionMenu(
            bar, width=180, variable=self.tmode_var,
            values=[
                "AI tự động (Gemini)",
                "Opencode tự động (đọc & dịch)",
                "Claude tự động (đọc & dịch)",
            ],
        ).pack(side="left")
        ctk.CTkLabel(bar, text="Model:").pack(side="left", padx=(12, 4))
        self.opencode_model_var = ctk.StringVar(value=config.OPENCODE_DEFAULT_MODEL)
        ctk.CTkOptionMenu(bar, width=160, variable=self.opencode_model_var,
                          values=config.OPENCODE_FREE_MODELS).pack(side="left")

        self.status_var = ctk.StringVar(value="Chưa nạp danh sách.")
        ctk.CTkLabel(self, textvariable=self.status_var, anchor="w").grid(
            row=1, column=0, sticky="ew", padx=8)

        self.table = ctk.CTkScrollableFrame(self, label_text="Danh sách video")
        self.table.grid(row=2, column=0, sticky="nsew", padx=4, pady=4)
        for i, w in enumerate((0, 1, 0, 0)):
            self.table.grid_columnconfigure(i, weight=w)

        self._poll()

    # --------------------------------------------------------------- load ----
    def _load(self):
        path = filedialog.askopenfilename(
            title="Chọn danh sách video",
            filetypes=[("JSON", "*.json"), ("Excel", "*.xlsx"), ("Tất cả", "*.*")],
        )
        if not path:
            return
        try:
            self.videos = self._read_list(path)
        except Exception as e:  # noqa: BLE001
            self.status_var.set(f"⚠ Không đọc được file: {e}")
            return
        self.path = path
        self._render_table()
        pending = sum(1 for v in self.videos if v.get("status") == "waiting")
        self.status_var.set(f"Đã nạp {len(self.videos)} video — {pending} chờ xử lý.")
        self.start_btn.configure(state="normal" if pending else "disabled")

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="Chọn các file video",
            filetypes=[("Video", "*.mp4 *.mkv *.webm *.mov *.avi"), ("Tất cả", "*.*")],
        )
        if paths:
            self.add_videos(list(paths))

    def add_videos(self, urls: list[str]):
        """Append YouTube URLs to the list (public API used by the YouTube tab).

        Keeps the existing JSON/Excel schema. If no list file is open yet, asks
        for a save path so _save_list() stays crash-safe across runs.
        """
        urls = [u.strip() for u in urls if u and u.strip()]
        if not urls:
            return
        if not self.path:
            path = filedialog.asksaveasfilename(
                title="Lưu danh sách video", defaultextension=".json",
                filetypes=[("JSON", "*.json")])
            if not path:
                return
            self.path = path
            if not self.videos:
                self.videos = []

        start_id = max((int(v.get("id", 0)) for v in self.videos), default=0)
        for i, url in enumerate(urls, start=start_id + 1):
            self.videos.append({"id": i, "video_url": url,
                                "voice_type": "male", "status": "waiting"})
        self._render_table()
        self._save_list()
        pending = sum(1 for v in self.videos if v.get("status") == "waiting")
        self.status_var.set(
            f"Đã thêm {len(urls)} video — tổng {len(self.videos)}, {pending} chờ xử lý.")
        if pending and not self._running:
            self.start_btn.configure(state="normal")

    def _open_youtube_dialog(self):
        win = ctk.CTkToplevel(self)
        win.title("Thêm video từ YouTube")
        win.geometry("620x360")
        win.after(120, lambda: (win.lift(), win.focus_force()))
        ctk.CTkLabel(
            win, justify="left", anchor="w",
            text="Dán URL YouTube (mỗi dòng 1 link). Link playlist sẽ tự bung "
                 "thành các video.").pack(fill="x", padx=16, pady=(14, 6))
        box = ctk.CTkTextbox(win, wrap="word")
        box.pack(fill="both", expand=True, padx=16, pady=(0, 8))
        status = ctk.CTkLabel(win, text="", anchor="w", text_color="gray70")
        status.pack(fill="x", padx=16)

        def add():
            lines = [ln.strip() for ln in box.get("1.0", "end").splitlines()
                     if ln.strip()]
            if not lines:
                status.configure(text="⚠ Chưa có URL nào.")
                return
            status.configure(text="Đang lấy thông tin playlist…")

            def work():
                from src.downloader import fetch_video_info
                collected: list[str] = []
                for ln in lines:
                    try:
                        info = fetch_video_info(ln)
                        if info.get("is_playlist"):
                            collected += [e["url"] for e in info.get("entries", [])
                                          if e.get("url")]
                        else:
                            collected.append(info.get("webpage_url") or ln)
                    except Exception:  # noqa: BLE001 - fall back to raw URL
                        collected.append(ln)
                self._post(lambda: (self.add_videos(collected), win.destroy()))

            threading.Thread(target=work, daemon=True).start()

        actions = ctk.CTkFrame(win, fg_color="transparent")
        actions.pack(fill="x", padx=16, pady=(4, 14))
        ctk.CTkButton(actions, text="Huỷ", fg_color="gray40", hover_color="gray30",
                      command=win.destroy).pack(side="right")
        ctk.CTkButton(actions, text="Thêm ▶", command=add).pack(side="right", padx=8)

    def _read_list(self, path: str) -> list[dict]:
        if path.lower().endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb.active
            videos = []
            for i, r in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                url = (r[0] if r else None)
                if not url:
                    continue
                status = (r[1] or "waiting") if len(r) > 1 else "waiting"
                videos.append({"id": i, "video_url": str(url).strip(),
                               "voice_type": "male", "status": str(status).strip().lower()})
            return videos
        with open(path, encoding="utf-8") as f:
            return json.load(f)

    def _save_list(self):
        if not self.path:
            return
        if self.path.lower().endswith(".xlsx"):
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            for v in self.videos:
                ws.append([v.get("video_url", ""), v.get("status", ""),
                           v.get("output_folder", "")])
            wb.save(self.path)
            return
        # Crash-safe JSON save (temp + replace), like batch_run_json.py.
        dir_name = os.path.dirname(os.path.abspath(self.path))
        fd, tmp = tempfile.mkstemp(suffix=".json", dir=dir_name)
        try:
            with os.fdopen(fd, "w", encoding="utf-8") as f:
                json.dump(self.videos, f, ensure_ascii=False, indent=2)
            os.replace(tmp, self.path)
        except Exception:
            if os.path.exists(tmp):
                os.remove(tmp)
            raise

    def _render_table(self):
        for w in self.table.winfo_children():
            w.destroy()
        self._rows.clear()
        headers = ["ID", "Video", "Giọng", "Trạng thái"]
        for c, h in enumerate(headers):
            ctk.CTkLabel(self.table, text=h, font=ctk.CTkFont(weight="bold")).grid(
                row=0, column=c, sticky="w", padx=6, pady=(2, 6))
        for r, v in enumerate(self.videos, start=1):
            ctk.CTkLabel(self.table, text=str(v.get("id", r))).grid(
                row=r, column=0, sticky="w", padx=6, pady=2)
            url = v.get("video_url", "")
            ctk.CTkLabel(self.table, text=(url[:60] + "…") if len(url) > 60 else url,
                         anchor="w").grid(row=r, column=1, sticky="w", padx=6, pady=2)
            ctk.CTkLabel(self.table, text=v.get("voice_type", "male")).grid(
                row=r, column=2, sticky="w", padx=6, pady=2)
            st = ctk.CTkLabel(self.table, text=v.get("status", "waiting"),
                              text_color=STATUS_COLORS.get(v.get("status", "waiting"), "gray60"))
            st.grid(row=r, column=3, sticky="w", padx=6, pady=2)
            self._rows[id(v)] = st

    def _set_status(self, video: dict, status: str):
        video["status"] = status
        label = self._rows.get(id(video))
        if label:
            label.configure(text=status, text_color=STATUS_COLORS.get(status, "gray60"))

    # -------------------------------------------------------------- run -------
    def _start(self):
        if self._running:
            return
        self._running = True
        self.start_btn.configure(state="disabled")
        threading.Thread(target=self._run_all, daemon=True).start()

    def _post(self, fn):
        self._ui_queue.put(fn)

    def _run_all(self):
        pending = [v for v in self.videos if v.get("status") == "waiting"]
        ok = fail = 0
        skip_video = self.skip_video_var.get()
        for i, video in enumerate(pending, start=1):
            src = video.get("video_url")
            is_file = bool(src) and os.path.exists(src)
            url = None if is_file else src
            file_path = src if is_file else None
            gender = "female" if video.get("voice_type") == "female" else "male"
            config.TTS_BACKEND_VI = config.TTS_BACKEND_VI  # honour current .env
            voice_id = config.vi_voice(gender)

            self._post(lambda v=video: (self._set_status(v, "processing"),
                                        self.status_var.set(
                                            f"[{i}/{len(pending)}] Đang xử lý: {src}")))
            self._save_list()
            start = time.time()
            try:
                bg_mode = self.bg_var.get()
                bg_duck_db = -12.0 if bg_mode == "duck" else 0.0
                phase1 = run_pipeline_vi(
                    url=url, file_path=file_path, source_lang=self.lang_var.get(),
                    voice_id=voice_id, skip_video=skip_video,
                    output_dir=_get_default_vi_output_dir(),
                    bg_mode=bg_mode, bg_duck_db=bg_duck_db,
                )
                work_dir = phase1.get("work_dir") if phase1.get("status") == "translate_pending" else phase1.get("output_dir")

                vi_path = os.path.join(work_dir, "transcript_vi.json")
                if not os.path.exists(vi_path):
                    mode = TRANSLATE_MODES.get(self.tmode_var.get(), "opencode")
                    source_lang = LANG_MAP.get(self.lang_var.get(), self.lang_var.get())
                    if mode == "ai":
                        with open(os.path.join(work_dir, "transcript_original.json"),
                                  encoding="utf-8") as f:
                            segments = json.load(f)
                        translated = translate_transcript(
                            segments=segments,
                            source_lang=source_lang,
                            api_key=config.GOOGLE_API_KEY,
                            model_id=config.CONTENT_MODEL_ID,
                        )
                        with open(vi_path, "w", encoding="utf-8") as f:
                            json.dump(translated, f, ensure_ascii=False, indent=2)
                    elif mode == "claude":
                        translate_via_claude_cli(
                            work_dir,
                            source_lang=source_lang,
                            model=(config.CLAUDE_MODEL_ID or None),
                        )
                    else:  # opencode (default)
                        translate_via_opencode_cli(
                            work_dir,
                            source_lang=source_lang,
                            model=self.opencode_model_var.get(),
                        )

                report = run_pipeline_vi(
                    url=url, file_path=file_path, source_lang=self.lang_var.get(),
                    voice_id=voice_id, skip_video=skip_video,
                    output_dir=_get_default_vi_output_dir(), resume_dir=work_dir,
                    bg_mode=bg_mode, bg_duck_db=bg_duck_db,
                )
                video["output_folder"] = report["session_id"]
                video["segments"] = report["total_segments"]
                video["duration_original"] = report["total_original_duration"]
                video["duration_vi"] = report["total_tts_duration"]
                video["processing_time"] = report["processing_time_seconds"]
                self._post(lambda v=video: self._set_status(v, "success"))
                ok += 1
            except Exception as e:  # noqa: BLE001
                video["error"] = str(e)[:200]
                self._post(lambda v=video: self._set_status(v, "failed"))
                fail += 1
            self._save_list()

        self._post(lambda: (self.status_var.set(
            f"Xong batch — {ok} thành công, {fail} lỗi."),
            self.start_btn.configure(state="normal")))
        self._running = False

    def _poll(self):
        try:
            while True:
                self._ui_queue.get_nowait()()
        except queue.Empty:
            pass
        self.after(200, self._poll)
